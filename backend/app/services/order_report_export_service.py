"""Exportação do relatório por OS (CSV, XLSX, PDF)."""

from __future__ import annotations

import base64
import csv
from collections import Counter
import functools
import io
import json
import os
import re
from datetime import UTC, datetime
from pathlib import Path
from typing import Literal

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from xhtml2pdf import pisa

from app.models.entities import ServiceOrder, ServiceOrderStatus
from app.repositories.robot_repository import RobotRepository
from app.schemas.service_order import MedicineReportLine, OrderReportItem
from app.services.historico_service import _data_conclusao_calendario_br

_TEMPLATES_DIR = Path(__file__).resolve().parent.parent / "templates"
_IMG_DIR = Path(__file__).resolve().parent.parent / "img"
_pdf_jinja_env: Environment | None = None


def _pdf_logos_files_exist() -> bool:
    return (_IMG_DIR / "SofttekLogo-4web.svg").is_file() and (_IMG_DIR / "logo_apsen.svg").is_file()


def _svg_file_to_png_data_uri(svg_path: Path) -> str | None:
    """Rasteriza SVG → PNG (mesma pilha do xhtml2pdf); data URI é mais fiável que <img src=.svg>."""
    try:
        from xhtml2pdf.xhtml2pdf_reportlab import PmlImage
    except ImportError:
        return None
    if not svg_path.is_file():
        return None
    try:
        raw = svg_path.read_bytes()
        pm = PmlImage(raw, src=svg_path.name)
        raster = pm.getDrawingRaster()
        if raster is None:
            return None
        b64 = base64.b64encode(raster.getvalue()).decode("ascii")
        return f"data:image/png;base64,{b64}"
    except Exception:
        return None


def _pdf_logo_image_srcs() -> tuple[str | None, str | None]:
    """Preferir PNG inline; fallback para ficheiros relativos + path no CreatePDF."""
    s_softtek = _svg_file_to_png_data_uri(_IMG_DIR / "SofttekLogo-4web.svg")
    s_apsen = _svg_file_to_png_data_uri(_IMG_DIR / "logo_apsen.svg")
    if s_softtek and s_apsen:
        return s_softtek, s_apsen
    if _pdf_logos_files_exist():
        return "SofttekLogo-4web.svg", "logo_apsen.svg"
    return None, None


def _pdf_img_base_path() -> str:
    """Diretório base para src relativos em <img> (pisa/xhtml2pdf)."""
    p = str(_IMG_DIR.resolve())
    if not p.endswith(os.sep):
        p += os.sep
    return p


def _pdf_medicamentos_por_classe(item: OrderReportItem) -> list[dict[str, str | int]]:
    """Contagem de medicamentos por classe (PDF): coluna classe + quantidade de itens."""
    counts: Counter[str] = Counter()
    for ml in item.medicine_lines:
        key = (ml.classe_remedio or "").strip() or "—"
        counts[key] += 1
    return [
        {"classe": k, "quantidade": counts[k]}
        for k in sorted(counts.keys(), key=lambda x: x.lower())
    ]


def _get_pdf_jinja_env() -> Environment:
    global _pdf_jinja_env
    if _pdf_jinja_env is None:
        _pdf_jinja_env = Environment(
            loader=FileSystemLoader(str(_TEMPLATES_DIR)),
            autoescape=select_autoescape(["html", "htm", "xml"]),
        )
    return _pdf_jinja_env


def _order_report_pdf_context(
    item: OrderReportItem,
    *,
    exportado_por: str,
    data_exportacao: str,
) -> dict:
    """Contexto para o template Jinja2 do PDF (A4, HTML)."""
    data_fmt = item.data.strftime("%d/%m/%Y") if item.data else "—"
    logo_softtek_src, logo_apsen_src = _pdf_logo_image_srcs()
    nome_sep = (item.separador_nome or "").strip() or "—"
    # Código do robô (ex.: RB-01) vem do cadastro; coincide com separador_codigo em order_to_report_item.
    rc = (item.separador_codigo or "").strip()
    robot_txt = (
        rc
        if rc
        else (str(item.robot_id) if item.robot_id is not None else "—")
    )
    return {
        "has_pdf_logos": logo_softtek_src is not None and logo_apsen_src is not None,
        "logo_softtek_src": logo_softtek_src or "",
        "logo_apsen_src": logo_apsen_src or "",
        "data_fmt": data_fmt,
        "os_code": item.os_code,
        "client_name": item.client_name or "",
        "nome_separador": nome_sep,
        "robot_codigo_txt": robot_txt,
        "quantidade_total": item.quantidade_total,
        "quantidade_separada": item.quantidade_separada,
        "porcentagem_conclusao": item.porcentagem_conclusao or "",
        "tempo_total_separacao": item.tempo_total_separacao or "",
        "tempo_liquido_separacao": item.tempo_liquido_separacao or "",
        "tempo_medio_por_remedio": item.tempo_medio_por_remedio or "",
        "numero_pausas": item.numero_pausas,
        "status_label": _situacao_label(item.situacao),
        "status_upper": _situacao_label_upper(item.situacao),
        "situacao": item.situacao,
        "erro_descricao": item.erro_descricao or "",
        "erro_codigo": item.erro_codigo or "",
        "exportado_por": exportado_por,
        "data_exportacao": data_exportacao,
        "medicamentos_por_classe": _pdf_medicamentos_por_classe(item),
    }


def _stamp_pdf_page_marks(pdf_bytes: bytes) -> bytes:
    """Carimba «Página N de M» no canto inferior direito de **cada** folha (o fluxo HTML só pinta na última)."""
    try:
        from pypdf import PdfReader, PdfWriter
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except ImportError:
        return pdf_bytes
    src_buf = io.BytesIO(pdf_bytes)
    try:
        reader = PdfReader(src_buf)
    except Exception:
        return pdf_bytes
    n = len(reader.pages)
    if n == 0:
        return pdf_bytes
    writer = PdfWriter()
    for idx in range(n):
        page = reader.pages[idx]
        media = page.mediabox
        w = float(media.width)
        h = float(media.height)
        packet = io.BytesIO()
        c = canvas.Canvas(packet, pagesize=(w, h))
        c.setFont("Helvetica", 8.5)
        c.setFillColorRGB(0.39, 0.45, 0.51)
        label = f"Página {idx + 1} de {n}"
        m = 12 * mm
        c.drawRightString(w - m, m, label)
        c.save()
        packet.seek(0)
        overlay = PdfReader(packet)
        page.merge_page(overlay.pages[0])
        writer.add_page(page)
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def _render_order_report_pdf_bytes(
    item: OrderReportItem,
    *,
    exportado_por: str,
    data_exportacao: str,
) -> bytes:
    env = _get_pdf_jinja_env()
    template = env.get_template("reports/os_report.html")
    html = template.render(
        **_order_report_pdf_context(
            item,
            exportado_por=exportado_por,
            data_exportacao=data_exportacao,
        )
    )
    buf = io.BytesIO()
    status = pisa.CreatePDF(
        src=html,
        dest=buf,
        encoding="utf-8",
        path=_pdf_img_base_path(),
    )
    if getattr(status, "err", 0):
        raise RuntimeError("Falha ao gerar PDF a partir do HTML.")
    raw = buf.getvalue()
    if not raw:
        raise RuntimeError("PDF gerado está vazio.")
    return _stamp_pdf_page_marks(raw)


_REPORT_HEADERS_BASE = [
    "Data",
    "Ordem de Serviço",
    "Cliente",
    "Quantidade Total",
    "Quantidade Separada",
    "Número de pausas",
    "Separador Utilizado",
    "Código do separador",
    "Remédio ID",
    "Remédio",
    "Tipo Remédio",
    "Classe Remédio",
    "Número",
    "Tempo Gasto",
    "Porcentagem de conclusão",
    "Tempo total da separação",
    "Tempo líquido da separação",
    "Tempo Médio por Remédio",
    "Status do Remédio",
    "Status da operação",
]
_REPORT_HEADERS_CANCEL = [
    "Descrição do Erro",
    "Código do erro",
]
_EXPORT_HEADERS = [
    "Exportado por",
    "Data da Exportação",
]


def _safe_filename_fragment(s: str) -> str:
    return re.sub(r"[^\w.\-]+", "_", s, flags=re.UNICODE).strip("._") or "OS"


def _normalize_utc(dt: datetime) -> datetime:
    if dt.tzinfo is None:
        return dt.replace(tzinfo=UTC)
    return dt


def _format_seconds_as_min_sec(sec: float | None) -> str:
    """Abaixo de 1 minuto: só segundos (ex.: 45 s). A partir de 1 min: «X min Y s»."""
    if sec is None or sec < 0:
        return ""
    total = int(round(sec))
    if total < 60:
        return f"{total} s"
    m, s = divmod(total, 60)
    return f"{m} min {s} s"


def _wall_seconds_total(order: ServiceOrder) -> float | None:
    """Tempo total (atribuição → encerramento), em segundos."""
    if order.status == ServiceOrderStatus.COMPLETED.value:
        if not order.assigned_at or not order.completed_at:
            return None
        a = _normalize_utc(order.assigned_at)
        c = _normalize_utc(order.completed_at)
        sec = (c - a).total_seconds()
        return sec if sec >= 0 else None
    if order.status == ServiceOrderStatus.CANCELLED.value:
        if order.cancelled_wall_seconds is not None:
            return float(order.cancelled_wall_seconds)
        return None
    return None


def _liquid_seconds(order: ServiceOrder) -> float | None:
    """Tempo efetivo (relógio menos pausas), em segundos; None se não for possível calcular."""
    wall = _wall_seconds_total(order)
    if wall is None:
        return None
    ps = order.total_pause_seconds
    if ps is None:
        if int(order.pause_count or 0) == 0:
            return float(wall)
        return None
    return max(0.0, float(wall) - float(max(0, int(ps))))


def _medicine_entries_from_order(order: ServiceOrder) -> list[tuple[str, str, str]]:
    """
    Lista (nome, tipo, classe) a partir de medicines_json.
    Itens em string só têm nome; objetos podem trazer tipo_remedio/tipo e classe_remedio/classe.
    """
    try:
        data = json.loads(order.medicines_json or "[]")
    except json.JSONDecodeError:
        return []
    if not isinstance(data, list):
        return []
    out: list[tuple[str, str, str]] = []
    for x in data:
        if isinstance(x, str):
            s = x.strip()
            if s:
                out.append((s, "", ""))
        elif isinstance(x, dict):
            nome = str(
                x.get("remedio") or x.get("nome") or x.get("name") or "",
            ).strip()
            tipo = str(
                x.get("tipo_remedio") or x.get("tipo") or "",
            ).strip()
            classe = str(
                x.get("classe_remedio") or x.get("classe") or "",
            ).strip()
            if nome:
                out.append((nome, tipo, classe))
        else:
            s = str(x).strip()
            if s:
                out.append((s, "", ""))
    return out


def _porcentagem_str(quantidade_total: int, quantidade_separada: int) -> str:
    if quantidade_total <= 0:
        return ""
    pct = min(100.0, 100.0 * float(quantidade_separada) / float(quantidade_total))
    return f"{pct:.1f}".replace(".", ",") + "%"


def _tempo_segundos_por_unidade(order: ServiceOrder) -> float | None:
    """Segundos por unidade (tempo líquido / unidades, quando disponível)."""
    if order.status == ServiceOrderStatus.COMPLETED.value:
        cu = int(order.completed_units or 0)
        if cu <= 0:
            return None
        liq = _liquid_seconds(order)
        if liq is None:
            return None
        return float(liq) / float(cu)

    if order.status == ServiceOrderStatus.CANCELLED.value:
        if order.cancelled_avg_seconds_per_unit is not None:
            return float(order.cancelled_avg_seconds_per_unit)
        u = int(order.cancelled_separated_units or 0)
        if u <= 0:
            return None
        liq = _liquid_seconds(order)
        if liq is None:
            return None
        return float(liq) / float(u)

    return None


def order_to_report_item(db: Session, order: ServiceOrder) -> OrderReportItem | None:
    """Monta o item do relatório por OS para uma ordem encerrada."""
    robots = RobotRepository(db)
    rids: set[int] = set()
    if order.completed_by_robot_id is not None:
        rids.add(order.completed_by_robot_id)
    if order.cancelled_by_robot_id is not None:
        rids.add(order.cancelled_by_robot_id)
    names = robots.get_names_by_ids(rids)
    codes = robots.get_codes_by_ids(rids)

    quantidade_total = max(0, int(order.expected_units or 0))

    if order.status == ServiceOrderStatus.COMPLETED.value:
        ca = order.completed_at
        if ca is None:
            return None
        d = _data_conclusao_calendario_br(ca)
        rid = order.completed_by_robot_id
        snap = (order.completed_by_robot_name or "").strip()
        sep_nome = snap or (names.get(rid) if rid is not None else None) or None
        quantidade_separada = max(
            0,
            int(order.completed_units) if order.completed_units is not None else 0,
        )
        situacao: Literal["concluida", "cancelada"] = "concluida"
        erro_desc = ""
        erro_cod = ""
    elif order.status == ServiceOrderStatus.CANCELLED.value:
        canc = order.cancelled_at
        if canc is None:
            return None
        d = _data_conclusao_calendario_br(canc)
        rid = order.cancelled_by_robot_id
        snap = (order.cancelled_by_robot_name or "").strip()
        sep_nome = snap or (names.get(rid) if rid is not None else None) or None
        quantidade_separada = max(0, int(order.cancelled_separated_units or 0))
        situacao = "cancelada"
        erro_desc = (order.cancel_error_description or "").strip()
        erro_cod = (order.cancel_error_code or "").strip()
    else:
        return None

    sep_codigo = (codes.get(rid) if rid is not None else None) or ""
    wall_s = _wall_seconds_total(order)
    liquid_s = _liquid_seconds(order)
    tempo_total_str = _format_seconds_as_min_sec(wall_s)
    tempo_liquido_str = _format_seconds_as_min_sec(liquid_s)
    tempo_s = _tempo_segundos_por_unidade(order)
    tempo_medio_str = _format_seconds_as_min_sec(tempo_s)

    entries = _medicine_entries_from_order(order)
    n_med = len(entries)
    medicine_lines: list[MedicineReportLine] = []
    if situacao == "concluida":
        if n_med > 0 and liquid_s is not None:
            per = liquid_s / float(n_med)
            for i, (name, tipo, classe) in enumerate(entries):
                med_num = i + 1
                medicine_lines.append(
                    MedicineReportLine(
                        remedio_id=str(med_num),
                        remedio=name,
                        tipo_remedio=tipo,
                        classe_remedio=classe,
                        numero=med_num,
                        situacao_coleta="concluida",
                        tempo_gasto=_format_seconds_as_min_sec(per),
                    )
                )
        elif n_med > 0:
            for i, (name, tipo, classe) in enumerate(entries):
                med_num = i + 1
                medicine_lines.append(
                    MedicineReportLine(
                        remedio_id=str(med_num),
                        remedio=name,
                        tipo_remedio=tipo,
                        classe_remedio=classe,
                        numero=med_num,
                        situacao_coleta="concluida",
                        tempo_gasto="",
                    )
                )
    else:
        # Cancelada: todas as linhas do pedido; índice < cancelled_separated_units = já separado; resto = ainda na fila.
        u_sep = min(max(0, int(order.cancelled_separated_units or 0)), n_med)
        tu = _tempo_segundos_por_unidade(order)
        tempo_sep = _format_seconds_as_min_sec(tu) if tu is not None else ""
        for i, (name, tipo, classe) in enumerate(entries):
            med_num = i + 1
            ja = i < u_sep
            sc: Literal["concluida", "cancelada"] = "concluida" if ja else "cancelada"
            medicine_lines.append(
                MedicineReportLine(
                    remedio_id=str(med_num),
                    remedio=name,
                    tipo_remedio=tipo,
                    classe_remedio=classe,
                    numero=med_num,
                    situacao_coleta=sc,
                    tempo_gasto=tempo_sep if ja else "",
                )
            )

    return OrderReportItem(
        id=order.id,
        os_code=order.os_code,
        client_name=order.client_name or "",
        data=d,
        separador_nome=sep_nome,
        quantidade_total=quantidade_total,
        quantidade_separada=quantidade_separada,
        tempo_medio_por_remedio=tempo_medio_str,
        unidades_totais=quantidade_separada,
        situacao=situacao,
        erro_descricao=erro_desc,
        erro_codigo=erro_cod,
        medicine_lines=medicine_lines,
        numero_pausas=max(0, int(order.pause_count or 0)),
        separador_codigo=str(sep_codigo).strip(),
        robot_id=rid,
        porcentagem_conclusao=_porcentagem_str(quantidade_total, quantidade_separada),
        tempo_total_separacao=tempo_total_str,
        tempo_liquido_separacao=tempo_liquido_str,
    )


def _situacao_label(s: Literal["concluida", "cancelada"]) -> str:
    return "Concluída" if s == "concluida" else "Cancelada"


def _situacao_label_upper(s: Literal["concluida", "cancelada"]) -> str:
    return "CONCLUÍDA" if s == "concluida" else "CANCELADA"


def _situacao_linha_remedio_label(s: Literal["concluida", "cancelada"]) -> str:
    """Status individual do remédio na linha (coleta concluída ou não)."""
    return "Concluído" if s == "concluida" else "Não concluído"


def _headers_full(include_cancel: bool) -> list[str]:
    h = list(_REPORT_HEADERS_BASE)
    if include_cancel:
        h += _REPORT_HEADERS_CANCEL
    h += _EXPORT_HEADERS
    return h


def _row_cells(
    item: OrderReportItem,
    med: MedicineReportLine | None,
    *,
    exportado_por: str,
    data_exportacao: str,
) -> list[str | int]:
    data_str = item.data.strftime("%d/%m/%Y") if item.data else ""
    if med is None:
        mid, mname, mtipo, mclasse, mnum, mt, mest = "", "", "", "", "", "", ""
    else:
        mid = med.remedio_id
        mname = med.remedio
        mtipo = med.tipo_remedio or ""
        mclasse = med.classe_remedio or ""
        mnum = med.numero
        mt = med.tempo_gasto
        mest = _situacao_linha_remedio_label(med.situacao_coleta)
    base: list[str | int] = [
        data_str,
        item.os_code,
        item.client_name or "",
        item.quantidade_total,
        item.quantidade_separada,
        item.numero_pausas,
        item.separador_nome or "",
        item.separador_codigo or "",
        mid,
        mname,
        mtipo,
        mclasse,
        mnum,
        mt,
        item.porcentagem_conclusao or "",
        item.tempo_total_separacao or "",
        item.tempo_liquido_separacao or "",
        item.tempo_medio_por_remedio or "",
        mest,
        _situacao_label(item.situacao),
    ]
    if item.situacao == "cancelada":
        base.extend([item.erro_descricao or "", item.erro_codigo or ""])
    base.extend([exportado_por, data_exportacao])
    return base


def _headers_batch() -> list[str]:
    """Cabeçalho único para export em lote (sempre inclui colunas de erro; vazias se OS concluída)."""
    return list(_REPORT_HEADERS_BASE) + list(_REPORT_HEADERS_CANCEL) + list(_EXPORT_HEADERS)


def _row_cells_batch(
    item: OrderReportItem,
    med: MedicineReportLine | None,
    *,
    exportado_por: str,
    data_exportacao: str,
) -> list[str | int]:
    """Uma linha no ficheiro de lote: mesmas colunas para todas as OS (erro vazio se concluída)."""
    data_str = item.data.strftime("%d/%m/%Y") if item.data else ""
    if med is None:
        mid, mname, mtipo, mclasse, mnum, mt, mest = "", "", "", "", "", "", ""
    else:
        mid = med.remedio_id
        mname = med.remedio
        mtipo = med.tipo_remedio or ""
        mclasse = med.classe_remedio or ""
        mnum = med.numero
        mt = med.tempo_gasto
        mest = _situacao_linha_remedio_label(med.situacao_coleta)
    base: list[str | int] = [
        data_str,
        item.os_code,
        item.client_name or "",
        item.quantidade_total,
        item.quantidade_separada,
        item.numero_pausas,
        item.separador_nome or "",
        item.separador_codigo or "",
        mid,
        mname,
        mtipo,
        mclasse,
        mnum,
        mt,
        item.porcentagem_conclusao or "",
        item.tempo_total_separacao or "",
        item.tempo_liquido_separacao or "",
        item.tempo_medio_por_remedio or "",
        mest,
        _situacao_label(item.situacao),
    ]
    err_d = (item.erro_descricao or "") if item.situacao == "cancelada" else ""
    err_c = (item.erro_codigo or "") if item.situacao == "cancelada" else ""
    base.extend([err_d, err_c, exportado_por, data_exportacao])
    return base


def _all_rows_batch(
    items: list[OrderReportItem],
    *,
    exportado_por: str,
    data_exportacao: str,
) -> list[list[str | int]]:
    rows: list[list[str | int]] = []
    kw = {"exportado_por": exportado_por, "data_exportacao": data_exportacao}
    for item in items:
        if item.medicine_lines:
            for ml in item.medicine_lines:
                rows.append(_row_cells_batch(item, ml, **kw))
        else:
            rows.append(_row_cells_batch(item, None, **kw))
    return rows


def export_batch_order_reports_bytes(
    items: list[OrderReportItem],
    fmt: Literal["csv", "xlsx"],
    *,
    exportado_por: str,
    data_exportacao: str,
    filename_base: str,
) -> tuple[bytes, str, str]:
    """Várias OS num único CSV ou XLSX (uma folha, um cabeçalho)."""
    headers = _headers_batch()
    rows = _all_rows_batch(
        items,
        exportado_por=exportado_por,
        data_exportacao=data_exportacao,
    )
    safe = _safe_filename_fragment(filename_base)

    if fmt == "csv":
        buf = io.StringIO()
        w = csv.writer(buf, delimiter=";", lineterminator="\n", quoting=csv.QUOTE_MINIMAL)
        w.writerow(headers)
        for r in rows:
            w.writerow(r)
        raw = "\ufeff" + buf.getvalue()
        return raw.encode("utf-8"), "text/csv; charset=utf-8", f"{safe}.csv"

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    bold = Font(bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold
    for r in rows:
        ws.append(r)
    ncols = len(headers)
    nrows = 1 + len(rows)
    last_c = get_column_letter(ncols)
    ws.auto_filter.ref = f"A1:{last_c}{nrows}"
    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    return (
        bio.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"{safe}.xlsx",
    )


def _report_headers_and_rows(
    item: OrderReportItem,
    *,
    exportado_por: str,
    data_exportacao: str,
) -> tuple[list[str], list[list[str | int]]]:
    include_cancel = item.situacao == "cancelada"
    headers = _headers_full(include_cancel)
    kw = {"exportado_por": exportado_por, "data_exportacao": data_exportacao}
    if item.medicine_lines:
        rows = [_row_cells(item, ml, **kw) for ml in item.medicine_lines]
    else:
        rows = [_row_cells(item, None, **kw)]
    return headers, rows


def export_order_report_bytes(
    item: OrderReportItem,
    fmt: Literal["csv", "xlsx", "pdf"],
    *,
    exportado_por: str = "",
    data_exportacao: str = "",
) -> tuple[bytes, str, str]:
    """
    Retorna (corpo, media_type, nome_arquivo_com_extensão).
    CSV não suporta linha congelada; no XLSX a 1.ª linha tem filtro e está fixa ao rolar.
    """
    base = _safe_filename_fragment(item.os_code)
    headers, rows = _report_headers_and_rows(
        item,
        exportado_por=exportado_por,
        data_exportacao=data_exportacao,
    )

    if fmt == "csv":
        buf = io.StringIO()
        w = csv.writer(buf, delimiter=";", lineterminator="\n", quoting=csv.QUOTE_MINIMAL)
        w.writerow(headers)
        for r in rows:
            w.writerow(r)
        raw = "\ufeff" + buf.getvalue()
        return raw.encode("utf-8"), "text/csv; charset=utf-8", f"relatorio_{base}.csv"

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"
        bold = Font(bold=True)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
        for r in rows:
            ws.append(r)
        ncols = len(headers)
        nrows = 1 + len(rows)
        last_c = get_column_letter(ncols)
        ws.auto_filter.ref = f"A1:{last_c}{nrows}"
        ws.freeze_panes = "A2"
        bio = io.BytesIO()
        wb.save(bio)
        return (
            bio.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"relatorio_{base}.xlsx",
        )

    # PDF — HTML (Jinja2) renderizado para A4 via xhtml2pdf
    pdf_bytes = _render_order_report_pdf_bytes(
        item,
        exportado_por=exportado_por,
        data_exportacao=data_exportacao,
    )
    return pdf_bytes, "application/pdf", f"relatorio_{base}.pdf"
