"""Catálogo de produtos reais carregado da planilha Excel do estoque.

Fornece `random_products(n)` que sorteia N produtos reais para popular OS de teste,
substituindo a geração aleatória por palavras-chave de princípios ativos.

Mapeamento de colunas:
  Excel "Código"                   → remedio_id
  Excel "SKU"                      → remedio
  Excel "Peso (g)"                 → peso_teorico
  Excel "Classe Terapêutica"       → tipo_remedio
  Excel "Finalidade / Especialidade" → classe_remedio
"""

from __future__ import annotations

import random
from pathlib import Path
from typing import Any

_CATALOG_PATH = (
    Path(__file__).resolve().parent.parent.parent / "data" / "Planilha_Produtos_Atualizada (1).xlsx"
)

_products: list[dict[str, Any]] | None = None


def _load() -> list[dict[str, Any]]:
    global _products
    if _products is not None:
        return _products

    catalog: list[dict[str, Any]] = []
    if _CATALOG_PATH.is_file():
        try:
            from openpyxl import load_workbook

            wb = load_workbook(_CATALOG_PATH, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            for row in rows[1:]:
                sku = row[1] if len(row) > 1 else None
                if sku is None:
                    continue
                peso_raw = row[2] if len(row) > 2 else None
                peso: float | None = None
                if peso_raw is not None:
                    try:
                        peso = float(peso_raw)
                    except (ValueError, TypeError):
                        peso = None
                catalog.append(
                    {
                        "remedio_id": str(row[0] or "").strip() if row[0] is not None else "",
                        "remedio": str(sku).strip(),
                        "peso_teorico": peso,
                        "tipo_remedio": str(row[3] or "").strip() if len(row) > 3 and row[3] else "",
                        "classe_remedio": str(row[4] or "").strip() if len(row) > 4 and row[4] else "",
                    }
                )
        except Exception:
            catalog = []

    _products = catalog
    return _products


def catalog_size() -> int:
    return len(_load())


def random_products(
    quantidade: int,
    *,
    rng: random.Random | None = None,
) -> list[dict[str, Any]]:
    """Sorteia `quantidade` produtos reais do catálogo (com reposição).

    Cada item é um dict com: remedio_id, remedio, tipo_remedio, classe_remedio, peso_teorico.
    Se o catálogo não estiver disponível, retorna produtos genéricos (fallback seguro).
    """
    catalog = _load()
    rg = rng or random
    n = max(0, int(quantidade))

    if not catalog:
        return [
            {
                "remedio_id": str(i + 1),
                "remedio": f"Produto {i + 1}",
                "tipo_remedio": "",
                "classe_remedio": "",
                "peso_teorico": None,
            }
            for i in range(n)
        ]

    return [dict(rg.choice(catalog)) for _ in range(n)]
